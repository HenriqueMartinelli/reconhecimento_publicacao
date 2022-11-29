from difflib import SequenceMatcher
import openpyxl
from tkinter.filedialog import askopenfilename

file = askopenfilename()
wb = openpyxl.load_workbook(file)
ws = wb.worksheets[1]

a = ws.cell(2,1).value.replace('\n',' ').strip()
for row in range(3, 50):
    b = ws.cell(row,1).value.replace('\n',' ').strip()
    diff = SequenceMatcher(a,b, autojunk=False)
